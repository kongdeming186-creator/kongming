import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

wb = openpyxl.Workbook()
ws = wb.active
ws.title = '工时评估表'

header_font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
content_font = Font(name='微软雅黑', size=9)
module_font = Font(name='微软雅黑', size=10, bold=True, color='1F4E79')
module_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
high_priority_fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
medium_priority_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
low_priority_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
complex_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
medium_c_fill = PatternFill(start_color='FFD93D', end_color='FFD93D', fill_type='solid')
simple_fill = PatternFill(start_color='6BCB77', end_color='6BCB77', fill_type='solid')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='B4C6E7'),
    right=Side(style='thin', color='B4C6E7'),
    top=Side(style='thin', color='B4C6E7'),
    bottom=Side(style='thin', color='B4C6E7')
)

headers = ['序号', '需求名称', '需求描述', '需求编码', '复杂度', '优先级', '产品设计', '前端开发', '后端开发', '测试', '合计', '备注']

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

workload_data = [
    {
        'module': '一、首页仪表盘',
        'items': [
            ('数据概览与图表', '展示平台核心数据指标，包括居民总数、标签覆盖数、预警数量、任务完成率等关键数据卡片；通过环形饼图展示居民类别分布，柱状图展示预警类型统计，堆叠柱状图展示任务完成情况，面积折线图展示月度数据趋势，并提供社区维度统计明细表', 'DASH-001', 'M', '高', 1, 2, 2, 1, '数据卡片+饼图+柱状图+折线图+明细表'),
        ]
    },
    {
        'module': '二、居民信息管理',
        'items': [
            ('居民档案与标签体系', '实现居民基础信息的全生命周期管理，包括居民列表分页展示、多条件筛选查询、居民详情页折叠面板展示、居民信息编辑功能；建立完整的居民标签体系，支持标签分类树形结构管理、标签徽章展示、标签添加移除操作，所有操作均留痕可追溯', 'RES-001', 'P', '高', 2, 3, 3, 2, '列表/详情/编辑/标签分类/操作留痕'),
            ('核心标签核实', '针对低保、残疾、老年、社保（4050人员）四类核心标签，对接第三方系统进行数据核验，确保居民实际享受的资格与标签一致。支持多字段核验逻辑、申报时间轴展示、子系统数据比对、异常状态标记等功能', 'RES-002', 'P', '高', 1, 2, 2, 1, '低保/残疾/老年/社保4类核心标签'),
            ('其他标签核实', '覆盖公租房、计生、重症、涉军、困境儿童、支农返汉、残疾学费补贴等其他类型标签的核实功能，每类标签根据业务特点配置不同的核验规则和数据来源，支持多标签组合查询和批量核实', 'RES-003', 'M', '中', 1, 1, 2, 1, '公租房/计生/重症/涉军/困境儿童等'),
        ]
    },
    {
        'module': '三、数据导入',
        'items': [
            ('批量导入与校验', '支持居民基础信息和标签信息的批量Excel导入，提供导入模板下载功能；导入过程中进行数据格式校验、重复数据检测、关联关系核验；支持数据分页预览、错误行高亮提示、导入结果统计反馈，成功数据自动入库，失败数据导出错误清单', 'IMP-001', 'P', '高', 1, 1, 2, 1, '居民/标签导入/预览/校验/结果反馈/模板'),
            ('导入记录管理', '记录所有数据导入操作的历史日志，包括导入时间、操作人员、导入文件、数据总量、成功数、失败数等信息，支持按时间范围和导入类型查询历史导入记录', 'IMP-002', 'S', '中', 0, 0, 1, 0, '历史记录查询'),
        ]
    },
    {
        'module': '四、预警管理',
        'items': [
            ('预警处置与规则', '实现预警信息的全流程管理，包括预警列表多条件筛选、预警详情完整展示、多方式预警处置（核实/修正/忽略）、处置状态流转；支持预警类型配置、预警规则引擎设置、标签互斥规则配置，灵活定义预警触发条件', 'WARN-001', 'P', '高', 1, 2, 3, 1, '列表/详情/处置/规则配置/互斥规则'),
            ('自动比对预警', '通过定时任务自动执行居民标签与第三方系统数据的比对，发现不一致自动生成预警记录；支持比对周期配置、预警等级设置、多种提醒方式（站内消息/短信/邮件），确保问题及时发现和处理', 'WARN-002', 'P', '高', 1, 1, 1, 1, '定时任务/自动比对/提醒通知'),
        ]
    },
    {
        'module': '五、任务管理',
        'items': [
            ('任务全生命周期', '实现任务从创建到完成的全流程管理，包括任务列表展示、任务详情查看、任务派发分配、任务状态追踪（待处理/处理中/已完成/已驳回）、任务状态机流转；支持预警一键转任务、批量任务下发功能，提高工作效率', 'TASK-001', 'P', '高', 1, 2, 2, 1, '列表/详情/派发/状态追踪/预警转任务'),
            ('任务统计', '提供任务完成情况统计卡片，展示任务总量、已完成数、待处理数、完成率等关键指标，支持按时间、人员、社区等维度统计任务完成情况', 'TASK-002', 'S', '中', 0, 1, 1, 0, '统计卡片/完成率'),
        ]
    },
    {
        'module': '六、统计报表',
        'items': [
            ('统计分析与报表', '提供多维度数据统计分析功能，包括居民类别统计、标签类型统计、预警类型统计、任务完成统计等；支持月度数据趋势分析、社区维度明细统计，所有报表均可导出为Excel格式，满足不同场景的数据需求', 'RPT-001', 'M', '中', 1, 2, 2, 1, '多维度统计/趋势/明细/导出'),
        ]
    },
    {
        'module': '七、系统对接',
        'items': [
            ('核心系统对接', '对接民政系统、残疾系统、公安户籍系统三个核心业务系统，获取居民低保信息、残疾证信息、户籍基本信息等核心数据，用于标签核实和数据比对；接口采用标准协议，确保数据安全和传输稳定', 'SYS-001', 'P', '高', 1, 1, 2, 1, '民政/残疾/公安户籍3个核心系统'),
            ('民生服务对接', '对接计生系统、养老平台、就业信息系统、工商信息系统等民生服务系统，获取计生信息、养老服务信息、就业失业信息、企业注册信息等数据，为相关标签核实提供数据支撑', 'SYS-002', 'P', '高', 1, 1, 2, 1, '计生/养老/就业/工商4个系统'),
        ]
    },
    {
        'module': '八、系统管理',
        'items': [
            ('用户权限与配置', '提供完整的系统管理功能，包括用户管理（新增/编辑/删除/启用禁用）、角色管理、权限分配（菜单权限/数据权限/操作权限）；支持社区网格树形结构管理、标签字典维护、操作日志审计，确保系统安全可控', 'MGT-001', 'P', '高', 1, 2, 2, 1, '用户/角色/权限/社区网格/标签字典/日志'),
        ]
    },
    {
        'module': '九、AI智能应用',
        'items': [
            ('AI智能助手', '集成AI大模型能力，提供智能对话交互界面；支持政策问答（基于政策知识库）、操作指引（功能使用说明）、自然语言数据查询、智能语义搜索、标签智能推荐、预警处置建议等AI功能，提升用户操作效率和智能化水平', 'AI-001', 'P', '中', 1, 2, 2, 1, '对话/问答/指引/查询/搜索/推荐/建议'),
        ]
    },
]

row_idx = 2
seq = 1
total_product = 0
total_frontend = 0
total_backend = 0
total_test = 0

for module in workload_data:
    ws.cell(row=row_idx, column=1, value=module['module'])
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=12)
    cell = ws.cell(row=row_idx, column=1)
    cell.font = module_font
    cell.fill = module_fill
    cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    cell.border = thin_border
    row_idx += 1

    module_product = 0
    module_frontend = 0
    module_backend = 0
    module_test = 0

    for item in module['items']:
        name, desc, code, complexity, priority, product, frontend, backend, test, note = item
        
        ws.cell(row=row_idx, column=1, value=seq)
        ws.cell(row=row_idx, column=2, value=name)
        ws.cell(row=row_idx, column=3, value=desc)
        ws.cell(row=row_idx, column=4, value=code)
        ws.cell(row=row_idx, column=5, value={'S': '简单', 'M': '中等', 'P': '复杂'}[complexity])
        ws.cell(row=row_idx, column=6, value=priority)
        ws.cell(row=row_idx, column=7, value=product)
        ws.cell(row=row_idx, column=8, value=frontend)
        ws.cell(row=row_idx, column=9, value=backend)
        ws.cell(row=row_idx, column=10, value=test)
        ws.cell(row=row_idx, column=11, value=product + frontend + backend + test)
        ws.cell(row=row_idx, column=12, value=note)
        
        module_product += product
        module_frontend += frontend
        module_backend += backend
        module_test += test
        
        for col in range(1, 13):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = content_font
            cell.border = thin_border
            cell.alignment = left_align if col in [2, 3, 12] else center_align
        
        complexity_color = {'S': simple_fill, 'M': medium_c_fill, 'P': complex_fill}
        ws.cell(row=row_idx, column=5).fill = complexity_color[complexity]
        priority_color = {'高': high_priority_fill, '中': medium_priority_fill, '低': low_priority_fill}
        ws.cell(row=row_idx, column=6).fill = priority_color[priority]
        
        seq += 1
        row_idx += 1
    
    ws.cell(row=row_idx, column=1, value='')
    ws.cell(row=row_idx, column=2, value=f'{module["module"]}小计')
    ws.cell(row=row_idx, column=3, value='')
    ws.cell(row=row_idx, column=4, value='')
    ws.cell(row=row_idx, column=5, value='')
    ws.cell(row=row_idx, column=6, value='')
    ws.cell(row=row_idx, column=7, value=module_product)
    ws.cell(row=row_idx, column=8, value=module_frontend)
    ws.cell(row=row_idx, column=9, value=module_backend)
    ws.cell(row=row_idx, column=10, value=module_test)
    ws.cell(row=row_idx, column=11, value=module_product + module_frontend + module_backend + module_test)
    ws.cell(row=row_idx, column=12, value='')
    
    for col in range(1, 13):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = Font(name='微软雅黑', size=9, bold=True)
        cell.border = thin_border
        cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        cell.alignment = center_align
    ws.cell(row=row_idx, column=2).alignment = left_align
    
    total_product += module_product
    total_frontend += module_frontend
    total_backend += module_backend
    total_test += module_test
    
    row_idx += 1

total = total_product + total_frontend + total_backend + total_test

ws.cell(row=row_idx, column=1, value='')
ws.cell(row=row_idx, column=2, value='总 计')
ws.cell(row=row_idx, column=3, value='')
ws.cell(row=row_idx, column=4, value='')
ws.cell(row=row_idx, column=5, value='')
ws.cell(row=row_idx, column=6, value='')
ws.cell(row=row_idx, column=7, value=total_product)
ws.cell(row=row_idx, column=8, value=total_frontend)
ws.cell(row=row_idx, column=9, value=total_backend)
ws.cell(row=row_idx, column=10, value=total_test)
ws.cell(row=row_idx, column=11, value=total)
ws.cell(row=row_idx, column=12, value=f'{seq-1}个功能点')

for col in range(1, 13):
    cell = ws.cell(row=row_idx, column=col)
    cell.font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
    cell.border = thin_border
    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    cell.alignment = center_align
ws.cell(row=row_idx, column=2).alignment = left_align

col_widths = [6, 20, 50, 12, 10, 10, 10, 10, 10, 10, 10, 35]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[chr(64 + i)].width = width

for i in range(1, row_idx + 1):
    ws.row_dimensions[i].height = 25

ws.freeze_panes = 'C2'

ws2 = wb.create_sheet('预算汇总')
ws2.column_dimensions['A'].width = 25
ws2.column_dimensions['B'].width = 15
ws2.column_dimensions['C'].width = 15
ws2.column_dimensions['D'].width = 15
ws2.column_dimensions['E'].width = 20

budget_data = [
    ('项目预算评估', '', '', '', ''),
    ('', '', '', '', ''),
    ('预算信息', '', '', '', ''),
    ('项目预算总额', '120,000', '元', '', ''),
    ('单价（元/人天）', '1,500', '元/人天', '', ''),
    ('可用总工时', '80', '人天', '', ''),
    ('', '', '', '', ''),
    ('工时分配', '', '', '', ''),
    ('角色', '工时(人天)', '金额(元)', '占比', ''),
    ('产品设计', total_product, total_product * 1500, f'{total_product/total*100:.1f}%'),
    ('前端开发', total_frontend, total_frontend * 1500, f'{total_frontend/total*100:.1f}%'),
    ('后端开发', total_backend, total_backend * 1500, f'{total_backend/total*100:.1f}%'),
    ('测试', total_test, total_test * 1500, f'{total_test/total*100:.1f}%'),
    ('总计', total, total * 1500, '100%'),
    ('', '', '', '', ''),
    ('预算使用', '', '', '', ''),
    ('实际工时', total, '人天', '', ''),
    ('实际费用', f'{total * 1500:,}', '元', '', ''),
    ('预算余额', f'{120000 - total * 1500:,}', '元', '', ''),
    ('', '', '', '', ''),
    ('工期估算', '', '', '', ''),
    ('团队配置', '人员', '工期(天)', '', ''),
    ('产品经理', 1, total_product, '', ''),
    ('前端开发', 2, total_frontend / 2, '', ''),
    ('后端开发', 2, total_backend / 2, '', ''),
    ('测试工程师', 1, total_test, '', ''),
    ('建议总工期', '', max(total_product, total_frontend / 2, total_backend / 2, total_test), '工作日', ''),
]

for i, row_data in enumerate(budget_data, 1):
    for j, value in enumerate(row_data, 1):
        cell = ws2.cell(row=i, column=j, value=value)
        cell.border = thin_border
        
        if i == 1:
            cell.font = Font(name='微软雅黑', size=14, bold=True)
        elif i == 3 or i == 7 or i == 15 or i == 22:
            cell.font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
            cell.fill = header_fill
        elif i == 8 or i == 23:
            cell.font = Font(name='微软雅黑', size=10, bold=True)
            cell.fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
        elif i == 9 or i == 24:
            cell.font = Font(name='微软雅黑', size=9, bold=True)
            cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        elif i == 14 or i == 20:
            cell.font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        else:
            cell.font = Font(name='微软雅黑', size=9)
            cell.alignment = center_align if j > 1 else left_align

ws3 = wb.create_sheet('说明')
ws3.column_dimensions['A'].width = 20
ws3.column_dimensions['B'].width = 55

legend_data = [
    ('复杂度说明', ''),
    ('简单(S)', '功能单一，页面数量少，无复杂业务逻辑'),
    ('中等(M)', '功能较复杂，涉及多个页面或简单业务逻辑'),
    ('复杂(P)', '功能复杂，涉及多系统对接、复杂业务规则'),
    ('', ''),
    ('优先级说明', ''),
    ('高', '核心业务功能，必须优先完成'),
    ('中', '重要功能，但可稍后实现'),
    ('低', '辅助功能，可最后实现'),
    ('', ''),
    ('工时说明', ''),
    ('单位', '人天（1人工作1天=1人天）'),
    ('单价', '1,500元/人天'),
    ('总预算', '120,000元 (80人天)'),
    ('', ''),
    ('注意事项', ''),
    ('1', '此工时为预估，实际开发中可根据情况调整'),
    ('2', '系统对接需与第三方协调，时间可能延长'),
    ('3', 'AI功能依赖AI接口稳定性'),
    ('4', '建议预留5-10%缓冲工时'),
    ('', ''),
    ('分阶段实施建议', ''),
    ('第一阶段', '首页仪表盘 + 居民信息管理核心（高优先级）'),
    ('第二阶段', '预警管理 + 任务管理 + 数据导入'),
    ('第三阶段', '统计报表 + 系统对接 + 系统管理'),
    ('第四阶段', 'AI智能应用（可作为后续迭代）'),
]

for i, (key, value) in enumerate(legend_data, 1):
    ws3.cell(row=i, column=1, value=key)
    ws3.cell(row=i, column=2, value=value)
    ws3.cell(row=i, column=1).border = thin_border
    ws3.cell(row=i, column=2).border = thin_border
    if key and value == '':
        ws3.cell(row=i, column=1).font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
        ws3.cell(row=i, column=1).fill = header_fill
    elif key and not key.isdigit() and not value == '':
        ws3.cell(row=i, column=1).font = Font(name='微软雅黑', size=9, bold=True)
    else:
        ws3.cell(row=i, column=1).font = Font(name='微软雅黑', size=9)

output_path = r'd:\原型\六角亭街道综合管理平台工时评估表.xlsx'
wb.save(output_path)

print('=' * 55)
print('  工时评估表生成成功！')
print('=' * 55)
print(f'文件路径: {output_path}')
print()
print('【预算约束】')
print(f'  项目预算: 120,000 元')
print(f'  单  价: 1,500 元/人天')
print(f'  可用工时: 80 人天')
print()
print('【工时分配】')
print(f'  产品设计: {total_product} 人天 ({total_product/total*100:.1f}%)')
print(f'  前端开发: {total_frontend} 人天 ({total_frontend/total*100:.1f}%)')
print(f'  后端开发: {total_backend} 人天 ({total_backend/total*100:.1f}%)')
print(f'  测    试: {total_test} 人天 ({total_test/total*100:.1f}%)')
print(f'  ─────────────────────')
print(f'  总    计: {total} 人天')
print()
print('【预算使用】')
print(f'  实际费用: {total * 1500:,} 元')
print(f'  预算余额: {120000 - total * 1500:,} 元')
print(f'  预算状态: {"✅ 预算充足" if total <= 80 else "❌ 超预算，需调整"}')
print()
print('【工期估算】（按最优并行计算）')
print(f'  产品设计: {total_product} 天 (1人)')
print(f'  前端开发: {total_frontend/2:.0f} 天 (2人)')
print(f'  后端开发: {total_backend/2:.0f} 天 (2人)')
print(f'  测    试: {total_test} 天 (1人)')
print(f'  ─────────────────────')
print(f'  建议总工期: {max(total_product, total_frontend/2, total_backend/2, total_test):.0f} 个工作日')
